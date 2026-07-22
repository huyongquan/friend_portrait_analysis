# -*- coding: utf-8 -*-
"""
朋友圈好友画像分析
用 qwen3.7-plus 多模态模型，对每位好友的 3 张朋友圈截图做内容理解，
输出用户画像（职业/兴趣/生活状态/活跃度等标签），生成 Excel。

用法:
    py -3.12 analyze.py --limit 5   # 只跑前5人(样本验证)
    py -3.12 analyze.py             # 跑全部(已完成的自动跳过, 可中断续跑)
    py -3.12 analyze.py --no-xlsx   # 只跑不生成 Excel
"""
import os, sys, json, base64, re, io, argparse, time, threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
import requests

# ========== 配置 ==========
API_KEY = "sk-sp-D.IYLYL.ZTMq.MEYCIQCuCCbG86UQvZAnbsSfI4ZjLwmYoIChQ14/ZmXuFDvnLAIhAImtBvJUPmH1jeEmMQ7M0EkGz143KxepRblI0LbeiCRD"
BASE_URL = "https://token-plan.cn-beijing.maas.aliyuncs.com/apps/anthropic"
MODEL = "qwen3.7-plus"

HERE = Path(__file__).parent
SCREENSHOT_DIR = Path(r"C:\Users\IDEA\Downloads\screenshot")
OUT_JSON = HERE / ".portrait_progress.json"   # 临时续跑文件(隐藏), 出Excel后自动删除
OUT_XLSX = HERE / "好友画像分析.xlsx"

MAX_W = 1000          # 压缩宽度上限(朋友圈文字小, 512读不清; 1000兼顾清晰度与payload)
JPEG_QUALITY = 85
MAX_WORKERS = 8
MAX_TOKENS = 4000

# 非好友的特殊账号(微信系统号/采集异常占位), 不做画像分析
EXCLUDE_NAMES = {"异常", "微信团队", "文件传输助手"}

# Excel 表头顺序 (微信昵称=文件名去掉-N后缀, 直接取, 不让模型读)
COLUMNS = [
    "微信昵称", "性别", "年龄段", "职业或行业", "兴趣标签",
    "生活状态", "活跃度", "朋友圈内容摘要", "潜在业务价值", "综合画像标签",
]
LIST_FIELDS = {"兴趣标签", "综合画像标签"}

PROMPT = """下面是某微信好友个人主页的 3 张朋友圈截图(依次为 -1/-2/-3)。
请仔细阅读图片中的朋友圈内容(文案、配图、发布时间、可见互动),对该好友做用户画像分析。

只输出一个 JSON 对象,不要任何解释、不要 markdown 包裹。字段如下:
{
  "性别": "男/女/不确定",
  "年龄段": "如 18-25 / 25-35 / 35-45 / 45-55 / 未知",
  "职业或行业": "如 保险代理人/宝妈/教师/个体经营/上班族/未知,尽量具体",
  "兴趣标签": ["从朋友圈内容归纳的兴趣,如 旅游/美食/亲子/健身/摄影/理财/养生,3-6个"],
  "生活状态": "如 已婚有娃/单身/创业者/退休/求学中",
  "活跃度": "高/中/低,后接简短依据,如 '中,近期更新较少'",
  "朋友圈内容摘要": "最近几条朋友圈的主题,30-60字",
  "潜在业务价值": "保险场景线索,如 家庭责任重/关注健康/有理财需求/为孩子规划/无明显线索",
  "综合画像标签": ["3-5个性格或画像标签,如 自律/爱晒娃/商务型/顾家/精致生活"]
}
注意: 读不清的字段填"未知"或空数组,不要编造。"""


# ========== 图片处理 ==========
def compress_to_b64(img_path):
    """读图 -> 压缩宽度 -> RGB -> JPEG base64。截图多为 RGBA,必须转RGB。"""
    from PIL import Image
    img = Image.open(img_path)
    w, h = img.size
    if w > MAX_W:
        h = int(h * MAX_W / w)
        img = img.resize((MAX_W, h), Image.LANCZOS)
    img = img.convert("RGB")
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=JPEG_QUALITY)
    return base64.b64encode(buf.getvalue()).decode()


# ========== 分组 ==========
def group_friends(screenshot_dir):
    """扫描目录,按 <微信昵称>-N.png 分组。返回 [(微信昵称, {1:p,2:p,3:p}), ...] 按名排序。"""
    pat = re.compile(r"^(.*)-(\d+)\.png$", re.IGNORECASE)
    groups = {}
    for f in sorted(screenshot_dir.iterdir()):
        name = f.name
        if name.startswith("."):
            continue
        m = pat.match(name)
        if not m:
            continue
        key = m.group(1)
        if key in EXCLUDE_NAMES:
            continue
        seq = int(m.group(2))
        groups.setdefault(key, {})[seq] = str(f)
    # 每人按 seq 排好, 转 list
    out = []
    for key in sorted(groups.keys(), key=lambda s: (len(s), s)):
        seqs = groups[key]
        out.append((key, seqs))
    return out


# ========== qwen 调用 (参考 match_materials.py:234 _qwen_post_text) ==========
def _qwen_post_text(payload, timeout=200):
    """发 qwen messages payload, 带 thinking 兜底重试(最多5次), 返回 text。"""
    headers = {
        "x-api-key": API_KEY,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }
    text = ""
    data = {}
    for attempt in range(5):
        try:
            r = requests.post(f"{BASE_URL}/v1/messages", json=payload, headers=headers, timeout=timeout)
            # 内容审核拒绝(400)直接快速失败, 不重试浪费时间
            if r.status_code == 400:
                try:
                    msg = r.json().get("message", r.text[:200])
                except Exception:
                    msg = r.text[:200]
                if "inappropriate" in msg.lower() or "InvalidParameter" in r.text:
                    raise RuntimeError(f"内容审核拒绝(图片含不当内容): {msg}")
            r.raise_for_status()
            data = r.json()
        except RuntimeError:
            raise
        except Exception as e:
            print(f"  ⚠️ qwen API 第{attempt+1}次异常: {e}")
            continue
        thinking = ""
        for c in data.get("content", []):
            if c.get("type") == "text":
                text = c["text"]
                break
            if c.get("type") == "thinking":
                thinking += c.get("thinking", "") or ""
        if text:
            break
        # 无 text: 从 thinking 提取 JSON 对象兜底
        m = re.search(r"\{.*\}", thinking, re.DOTALL)
        if m:
            text = m.group(0)
            print(f"  ⚠️ 第{attempt+1}次无 text, 从 thinking 兜底提取到 JSON")
            break
        print(f"  ⚠️ qwen API 第{attempt+1}次无 text(仅thinking), 重试...")
    if not text:
        raise RuntimeError(f"qwen 无 text 回复(重试5次仍失败): {json.dumps(data, ensure_ascii=False)[:300]}")
    return text


def extract_json_object(text):
    """从模型回复提取 JSON 对象。处理 ```json 包裹 / 首末花括号。"""
    m = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, re.DOTALL)
    if m:
        text = m.group(1)
    else:
        start = text.find("{")
        end = text.rfind("}")
        if start >= 0 and end > start:
            text = text[start:end + 1]
    return json.loads(text)


def dedup_seqs(seqs):
    """同一好友多张截图若内容相同(字节md5一致), 只保留1张。
    seqs: {seq: path}。返回去重后的 {seq: path}(按seq升序, 内容相同的只留最小seq)。
    截图工具对空朋友圈/同一页截出的png字节级一致, md5判定快且零误判。"""
    import hashlib
    seen, out = {}, {}
    for seq in sorted(seqs.keys()):
        p = seqs[seq]
        try:
            h = hashlib.md5(Path(p).read_bytes()).hexdigest()
        except Exception:
            h = str(seq)  # 读不出就当作唯一, 不去重
        if h not in seen:
            seen[h] = seq
            out[seq] = p
    return out


def analyze_one(name, seqs):
    """分析一位好友。返回结果 dict(含 微信昵称)。"""
    t0 = time.time()
    # 先去重(内容相同的截图只留1张), 再按 seq 升序取图
    uniq = dedup_seqs(seqs)
    ordered = [uniq[k] for k in sorted(uniq.keys())]
    content = []
    for p in ordered:
        try:
            b64 = compress_to_b64(p)
        except Exception as e:
            return {"微信昵称": name, "error": f"图片处理失败: {e}"}
        content.append({
            "type": "image",
            "source": {"type": "base64", "media_type": "image/jpeg", "data": b64},
        })
    content.append({"type": "text", "text": PROMPT})
    payload = {"model": MODEL, "max_tokens": MAX_TOKENS, "messages": [{"role": "user", "content": content}]}
    text = _qwen_post_text(payload)
    obj = extract_json_object(text)
    obj["微信昵称"] = name
    obj["_耗时秒"] = round(time.time() - t0, 1)
    return obj


# ========== 增量保存 ==========
_save_lock = threading.Lock()

def load_done(out_json=None):
    """读已有结果,返回 {微信昵称: record}。out_json 缺省用模块全局 OUT_JSON。"""
    p = Path(out_json) if out_json else OUT_JSON
    if not p.exists():
        return {}
    try:
        arr = json.loads(p.read_text(encoding="utf-8"))
        return {r.get("微信昵称"): r for r in arr if r.get("微信昵称")}
    except Exception:
        return {}

def save_all(done_map, out_json=None):
    """全量写回(线程安全)。out_json 缺省用模块全局 OUT_JSON。"""
    p = Path(out_json) if out_json else OUT_JSON
    with _save_lock:
        arr = list(done_map.values())
        p.write_text(json.dumps(arr, ensure_ascii=False, indent=2), encoding="utf-8")


# ========== Excel ==========
def to_cell(val):
    """列表字段转 '、' 连接字符串; None/缺失转空串。"""
    if val is None:
        return ""
    if isinstance(val, list):
        return "、".join(str(x) for x in val if x)
    return str(val)

def write_xlsx(done_map, out_xlsx=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    out_path = Path(out_xlsx) if out_xlsx else OUT_XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = "好友画像"
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    rows = sorted(done_map.values(), key=lambda r: r.get("微信昵称", ""))
    n = 0
    for r in rows:
        name = r.get("微信昵称", "")
        if r.get("error"):
            ws.append([name, f"[失败] {r['error']}"] + [""] * (len(COLUMNS) - 2))
            n += 1
            continue
        ws.append([to_cell(r.get(c, "")) for c in COLUMNS])
        n += 1

    # 列宽自适应(粗略按中文2字符宽)
    for i, col in enumerate(COLUMNS, 1):
        max_len = len(col)
        for row in ws.iter_rows(min_row=2, min_col=i, max_col=i, values_only=True):
            v = row[0]
            if v is None:
                continue
            s = str(v)
            w = sum(2 if ord(ch) > 127 else 1 for ch in s)
            if w > max_len:
                max_len = w
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 60)

    try:
        wb.save(out_path)
        print(f"\n✅ Excel 已生成: {out_path}  共 {n} 行")
    except PermissionError:
        # 文件被占用(Excel 正开着), 改写到带时间戳的副本
        import time as _t
        alt = out_path.with_name(out_path.stem + f"_{int(_t.time())}" + out_path.suffix)
        wb.save(alt)
        print(f"\n⚠️ {out_path.name} 被占用, 已写到副本: {alt}  共 {n} 行")
    return n


def finish(done_map, out_xlsx=None, out_json=None):
    """生成 Excel, 保留临时续跑文件(.portrait_progress.json)以便后续续跑。"""
    n = write_xlsx(done_map, out_xlsx=out_xlsx)
    nf = sum(1 for r in done_map.values() if r.get("error"))
    jp = Path(out_json).name if out_json else OUT_JSON.name
    print(f"💾 结果已存 {jp} (保留, 重跑同命令可续跑跳过已完成的)")
    if nf:
        print(f"⚠️ 有 {nf} 人分析失败(如内容审核拒绝), 已在 Excel 标记。")
    return n, nf


def run_analysis(shot_dir, out_xlsx, out_json, on_progress=None,
                 limit=0, no_xlsx=False, cancel_check=None):
    """对 shot_dir 跑画像分析, 结果写 out_json(续跑) 和 out_xlsx(Excel)。
    on_progress(event, data): 进度回调(start/ok/fail/done), 供 web 更新任务状态。
    cancel_check(): 返回 True 则中止(供 web /cancel)。
    返回 (done_count, fail_count)。"""
    shot_dir = Path(shot_dir)
    friends = group_friends(shot_dir)
    print(f"扫描到 {len(friends)} 位好友 (目录 {shot_dir})")

    done_map = load_done(out_json)
    if done_map:
        print(f"已有结果 {len(done_map)} 条, 将跳过")

    todo = [(n, s) for n, s in friends if n not in done_map]
    if limit:
        todo = todo[:limit]
    print(f"本次待跑 {len(todo)} 人, 并发 {MAX_WORKERS}\n")

    if on_progress:
        on_progress("start", {"total": len(todo), "have_done": len(done_map)})

    if not todo:
        if not no_xlsx:
            finish(done_map, out_xlsx=out_xlsx, out_json=out_json)
        if on_progress:
            on_progress("done", {"done": 0, "fail": 0, "skipped": len(done_map)})
        return 0, 0

    done_count = 0
    fail_count = 0
    t_start = time.time()

    def worker(item):
        name, seqs = item
        try:
            return analyze_one(name, seqs)
        except Exception as e:
            return {"微信昵称": name, "error": str(e)[:200]}

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futs = {ex.submit(worker, item): item for item in todo}
        for fut in as_completed(futs):
            item = futs[fut]
            # 取消检测
            if cancel_check and cancel_check():
                for f in futs:
                    f.cancel()
                print("⚠️ 任务已取消")
                break
            try:
                res = fut.result()
            except Exception as e:
                res = {"微信昵称": item[0], "error": str(e)[:200]}
            done_map[res["微信昵称"]] = res
            save_all(done_map, out_json=out_json)
            elapsed = time.time() - t_start
            if res.get("error"):
                fail_count += 1
                print(f"[失败] {res['微信昵称']}: {res['error']}")
                if on_progress:
                    on_progress("fail", {"name": res["微信昵称"], "error": res["error"],
                                         "done": done_count, "fail": fail_count,
                                         "total": len(todo), "elapsed": elapsed})
            else:
                done_count += 1
                dur = res.get("_耗时秒", 0)
                job = res.get("职业或行业", "")
                tags = "、".join(res.get("综合画像标签", []) or [])
                print(f"[OK {done_count}/{len(todo)}] {res['微信昵称']}  职业={job}  标签={tags}  ({dur}s)")
                if on_progress:
                    on_progress("ok", {"name": res["微信昵称"], "job": job, "tags": tags,
                                        "dur": dur, "done": done_count, "fail": fail_count,
                                        "total": len(todo), "elapsed": elapsed})

    print(f"\n本轮完成: 成功 {done_count}, 失败 {fail_count}")
    if not no_xlsx:
        finish(done_map, out_xlsx=out_xlsx, out_json=out_json)
    if on_progress:
        on_progress("done", {"done": done_count, "fail": fail_count, "total": len(todo)})
    return done_count, fail_count


# ========== 主流程 ==========
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", default=str(SCREENSHOT_DIR), help=f"截图目录(默认 {SCREENSHOT_DIR})")
    ap.add_argument("--limit", type=int, default=0, help="只跑前N人(0=全部)")
    ap.add_argument("--no-xlsx", action="store_true", help="不生成Excel")
    ap.add_argument("--only-xlsx", action="store_true", help="跳过分析,直接用已有结果生成Excel")
    args = ap.parse_args()

    shot_dir = Path(args.dir)
    if not shot_dir.exists():
        print(f"❌ 截图目录不存在: {shot_dir}")
        return

    # Excel 和临时续跑文件都落在输入的截图目录下(输入哪个目录, 结果就在哪)
    out_json = shot_dir / ".portrait_progress.json"
    out_xlsx = shot_dir / "好友画像分析.xlsx"

    if args.only_xlsx:
        done_map = load_done(out_json)
        print(f"已有结果 {len(done_map)} 条, 直接生成 Excel")
        finish(done_map, out_xlsx=out_xlsx, out_json=out_json)
        return

    run_analysis(shot_dir, out_xlsx, out_json,
                 limit=args.limit, no_xlsx=args.no_xlsx)


if __name__ == "__main__":
    main()
